import pandas as pd
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def safe_json_load(x):
    """Safely parse the Purview AuditData JSON string."""
    try:
        return json.loads(str(x))
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}

def toggle_options():
    interactive_frame.pack_forget()
    non_interactive_frame.pack_forget()
    sui_frame.pack_forget()
    suni_frame.pack_forget()
    purview_frame.pack_forget()
    color_sort_frame.pack_forget()
    
    selection = log_type_var.get()
    if selection == "Interactive":
        interactive_frame.pack(fill="x")
    elif selection == "NonInteractive":
        non_interactive_frame.pack(fill="x")
    elif selection == "SUI":
        sui_frame.pack(fill="x")
    elif selection == "SUNI":
        suni_frame.pack(fill="x")
    elif selection == "Purview":
        purview_frame.pack(fill="x")
    elif selection == "ColorSort":
        color_sort_frame.pack(fill="x")

def run_check():
    file_path = file_path_var.get()
    if not file_path:
        messagebox.showwarning("Input Error", "Please select a file first.")
        return

    log_type = log_type_var.get()
    
    try:
        if file_path.lower().endswith('.csv'):
            headers = pd.read_csv(file_path, nrows=0, low_memory=False).columns
        else:
            headers = pd.read_excel(file_path, nrows=0).columns
    except Exception as e:
        messagebox.showerror("File Error", f"Could not read the file:\n{e}")
        return

    col_lower = [col.strip().lower() for col in headers]
    
    if log_type == "ColorSort":
        if 'session id' not in col_lower:
            messagebox.showerror("Validation Error", "The file must contain a 'Session ID' column for this feature.")
            return
    else:
        ni_indicators = ['client credential type', 'managed identity type', 'federated token id']
        purview_indicators = ['auditdata', 'recordid', 'recordtype']
        
        is_likely_ni = any(ind in col_lower for ind in ni_indicators)
        is_likely_purview = any(ind in col_lower for ind in purview_indicators)

        if log_type == "Purview" and not is_likely_purview:
            if not messagebox.askyesno("Dataset Mismatch", "This file is missing 'AuditData' or 'RecordType' and does not appear to be a Purview log.\n\nAre you sure you want to proceed?"):
                return
        elif log_type in ["Interactive", "SUI"] and is_likely_ni:
            if not messagebox.askyesno("Dataset Mismatch", "This file appears to contain populated Non-Interactive columns.\n\nAre you sure you want to run Interactive checks?"):
                return

    params = {
        'file_path': file_path,
        'log_type': log_type,
        'interactive': {'ips': ip_filter_var.get(), 'travel': travel_filter_var.get(), 'errors': errors_filter_var.get(), 'iocs': iocs_filter_var.get()},
        'ni': {'travel': ni_travel_var.get(), 'agents': ni_agents_var.get(), 'unmanaged': ni_unmanaged_var.get(), 'refresh': ni_refresh_var.get(), 'legacy': ni_legacy_var.get()},
        'sui': {'travel': sui_travel_var.get(), 'errors': sui_errors_var.get(), 'iocs': sui_iocs_var.get()},
        'suni': {'travel': suni_travel_var.get(), 'agents': suni_agents_var.get(), 'unmanaged': suni_unmanaged_var.get(), 'refresh': suni_refresh_var.get(), 'legacy': suni_legacy_var.get()},
        'purview': {'inbox': purview_inbox_var.get(), 'mass_ops': purview_mass_op_var.get(), 'edisc': purview_edisc_var.get(), 'mailbox': purview_mailbox_var.get()}
    }

    status_label.config(text=f"Processing...\nFile: {os.path.basename(file_path)}")
    run_btn.config(state="disabled")
    threading.Thread(target=process_analysis, args=(params,), daemon=True).start()

def generate_safe_color():
    while True:
        r = random.randint(180, 255)
        g = random.randint(180, 255)
        b = random.randint(200, 255) 
        
        if r > 220 and g < 200 and b < 200: continue
        if g > 220 and r < 200 and b < 200: continue
        if abs(r-g) < 10 and abs(g-b) < 10 and r > 230: continue
        
        return f"{r:02X}{g:02X}{b:02X}"

def process_color_sort(params):
    try:
        file_path = params['file_path']
        output_dir = os.path.dirname(file_path)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = os.path.join(output_dir, f"{base_name}_ColorSorted.xlsx")

        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path, low_memory=False)
            wb = Workbook()
            ws = wb.active
            ws.append(list(df.columns))
            for row in df.itertuples(index=False, name=None):
                ws.append(row)
        else:
            wb = load_workbook(file_path)
            ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        session_col_idx = None
        for idx, h in enumerate(headers):
            if h and str(h).strip().lower() == 'session id':
                session_col_idx = idx + 1
                break
                
        if session_col_idx is None:
            app.after(0, lambda: update_ui_error("Column 'Session ID' not found in the file."))
            return

        session_colors = {}
        
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=session_col_idx)
            session_id = cell.value
            
            if session_id:
                if session_id not in session_colors:
                    session_colors[session_id] = generate_safe_color()
                
                fill_color = session_colors[session_id]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        wb.save(output_filename)
        app.after(0, lambda: update_ui_success(output_filename, True, is_color_sort=True))

    except Exception as e:
        app.after(0, lambda: update_ui_error(f"Failed to process Color Sort:\n{e}"))

def process_analysis(params):
    if params['log_type'] == "ColorSort":
        process_color_sort(params)
        return

    try:
        file_path = params['file_path']
        log_type = params['log_type']
        
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path, low_memory=False)
        else:
            df = pd.read_excel(file_path)
            
        df.columns = df.columns.str.strip()
        
        findings_found = False
        summary = {"Critical": [], "Suspicious": [], "Info": []}
        details = []
        
        if log_type in ["Interactive", "SUI"]:
            findings_found, summary, details = analyze_interactive(df, log_type, params)
        elif log_type in ["NonInteractive", "SUNI"]:
            findings_found, summary, details = analyze_non_interactive(df, log_type, params)
        elif log_type == "Purview":
            findings_found, summary, details = analyze_purview(df, log_type, params)
            
        if summary.get("ERROR"):
            app.after(0, lambda: update_ui_error(summary["ERROR"]))
            return

        write_report(params['file_path'], log_type, findings_found, summary, details)
        
    except Exception as e:
        app.after(0, lambda: update_ui_error(f"Failed to analyze data:\n{e}"))

def write_report(file_path, log_type, findings_found, summary, details):
    try:
        output_dir = os.path.dirname(file_path)
        output_filename = os.path.join(output_dir, f"{log_type}_ThreatHunt_QuickReport.txt")
        
        with open(output_filename, 'w', encoding='utf-8') as f:
            f.write(f"SOC ENTRA ID / PURVIEW THREAT HUNT REPORT ({log_type.upper()} LOGS)\n")
            f.write("=" * 70 + "\n\n")

            if not findings_found:
                f.write("[+] All clear! No suspicious activity found based on selected criteria.\n")
            else:
                f.write("--- EXECUTIVE SUMMARY ---\n\n")
                for category in ["Critical", "Suspicious", "Info"]:
                    if summary[category]:
                        f.write(f"[{category.upper()}]\n")
                        for item in summary[category]:
                            f.write(f"  * {item}\n")
                        f.write("\n")
                f.write("=" * 70 + "\n\n")
                f.write("--- DETAILED FINDINGS ---\n\n")
                for line in details:
                    f.write(line)
                    
        app.after(0, lambda: update_ui_success(output_filename, findings_found))
    except Exception as e:
        app.after(0, lambda: update_ui_error(f"Failed to save report:\n{e}"))

def update_ui_success(output_filename, findings_found, is_color_sort=False):
    status_label.config(text=f"Status: Done! Saved to {os.path.basename(output_filename)}")
    run_btn.config(state="normal")
    if is_color_sort:
        messagebox.showinfo("Color Sort Complete", f"File successfully processed and saved to:\n{os.path.abspath(output_filename)}")
    elif not findings_found:
        messagebox.showinfo("Analysis Complete", "No suspicious findings. All clear!")
    else:
        messagebox.showwarning("Findings Detected", f"Suspicious activity found based on your parameters.\n\nReport saved to:\n{os.path.abspath(output_filename)}")

def update_ui_error(msg):
    status_label.config(text="Status: Waiting for input")
    run_btn.config(state="normal")
    messagebox.showerror("Error", msg)

def apply_vectorized_location(df):
    if 'Location' in df.columns:
        split_loc = df['Location'].astype(str).str.split(',')
        df['State'] = split_loc.str[1].str.strip()
        df['State'] = df['State'].fillna(df['Location'].astype(str))
        df['Country'] = split_loc.str[-1].str.strip()
        df.loc[df['Location'].isna(), ['State', 'Country']] = None, "Unknown"

def analyze_interactive(df, log_type, params):
    findings_found = False
    summary = {"Critical": [], "Suspicious": [], "Info": []}
    details = []
    
    cfg = params['interactive'] if log_type == "Interactive" else params['sui']
    
    req_cols = ['Session ID', 'User'] if log_type == "Interactive" else ['User']
    if cfg.get('ips', False): req_cols.append('IP address')
    if cfg['travel']: req_cols.extend(['Date (UTC)', 'Location'])
    if cfg['errors']: req_cols.append('Sign-in error code')
    if cfg['iocs']: req_cols.extend(['Client app', 'User agent'])

    missing = [col for col in req_cols if col not in df.columns]
    if missing:
        summary["ERROR"] = f"Missing required columns:\n{', '.join(missing)}"
        return False, summary, details

    if log_type == "SUI" and 'Location' in df.columns:
        findings_found = True
        loc_stats = df.dropna(subset=['Location']).groupby('Location').agg(
            count=('User', 'size'),
            unique_ips=('IP address', 'nunique') if 'IP address' in df.columns else ('User', lambda x: 0),
            ips=('IP address', lambda x: ', '.join(set(x.dropna().astype(str)))) if 'IP address' in df.columns else ('User', lambda x: 'Unknown')
        ).sort_values(by='count', ascending=False)
        
        summary["Info"].append("SUI Sign-in Locations Profile generated.")
        details.append("--- SIGN-IN LOCATIONS SUMMARY (Most to Least Common) ---\n")
        for loc, row in loc_stats.iterrows():
            ip_warn = " [!] MULTIPLE IPs DETECTED" if row['unique_ips'] > 1 else ""
            details.append(f"Location: {loc} | Count: {row['count']} | Unique IPs: {row['unique_ips']}{ip_warn}\n")
            if row['unique_ips'] > 1:
                details.append(f"  IP Addresses: {row['ips']}\n")
        details.append("\n" + "-" * 50 + "\n")

    if cfg.get('ips', False) and log_type == "Interactive":
        session_stats = df.groupby('Session ID').agg(unique_users=('User', 'nunique'), unique_ips=('IP address', 'nunique'))
        flagged = session_stats[(session_stats['unique_users'] > 1) & (session_stats['unique_ips'] > 1)]
        if not flagged.empty:
            findings_found = True
            summary["Suspicious"].append(f"Shared Sessions: {len(flagged)} instance(s) found.")
            details.append(f"[!] WARNING: Found {len(flagged)} Session ID(s) shared by multiple users AND multiple IP addresses.\n")
            details.append("-" * 50 + "\n")
            results = df[df['Session ID'].isin(flagged.index)].groupby('Session ID').agg(users=('User', 'unique'), ips=('IP address', 'unique')).reset_index()
            for _, row in results.iterrows():
                details.append(f"Session ID: {row['Session ID']}\n  Users: {', '.join(map(str, row['users']))}\n  IPs: {', '.join(map(str, row['ips']))}\n\n")

    if cfg['travel']:
        df['Date (UTC)'] = pd.to_datetime(df['Date (UTC)'], errors='coerce')
        apply_vectorized_location(df)
        
        if 'Sign-in error code' in df.columns:
            df['Clean_Error_Travel'] = df['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
            valid_codes = ['0', 'None', 'nan', '', 'NaN', '50140', '50158', '50074', '50076', '50079', '50125']
            travel_df = df[df['Clean_Error_Travel'].isin(valid_codes)].copy()
        else:
            travel_df = df.copy()

        travel_df = travel_df.sort_values(['User', 'Date (UTC)']).dropna(subset=['Date (UTC)', 'State'])
        travel_df['Prev_State'] = travel_df.groupby('User')['State'].shift(1)
        travel_df['Prev_Location'] = travel_df.groupby('User')['Location'].shift(1)
        travel_df['Prev_Time'] = travel_df.groupby('User')['Date (UTC)'].shift(1)
        travel_df['Prev_IP'] = travel_df.groupby('User')['IP address'].shift(1) if 'IP address' in travel_df.columns else "Unknown"
        travel_df['Time_Diff_Mins'] = (travel_df['Date (UTC)'] - travel_df['Prev_Time']).dt.total_seconds() / 60.0
        
        imp_travel = travel_df[
            (travel_df['State'] != travel_df['Prev_State']) & 
            (travel_df['Prev_State'].notna()) & 
            (travel_df['Time_Diff_Mins'] < 30.0) &
            (travel_df['Time_Diff_Mins'] >= 0)
        ]

        if not imp_travel.empty:
            findings_found = True
            summary["Critical"].append(f"Impossible Travel (<30 mins): {len(imp_travel)} instance(s) detected.")
            
            travel_details = []
            travel_details.append(f"[!] WARNING: Found {len(imp_travel)} instance(s) of Impossible Travel (<30 mins between states).\n")
            travel_details.append("-" * 50 + "\n")
            
            high_risk_users = set()
            user_blocks = []
            
            for user, group in imp_travel.groupby('User'):
                loc_ips = set()
                states = set()
                
                for _, row in group.iterrows():
                    cur_loc = row['Location'] if pd.notna(row['Location']) else 'Unknown'
                    cur_ip = row.get('IP address', 'Unknown')
                    if pd.isna(cur_ip): cur_ip = 'Unknown'
                    
                    prev_loc = row['Prev_Location'] if pd.notna(row['Prev_Location']) else 'Unknown'
                    prev_ip = row.get('Prev_IP', 'Unknown')
                    if pd.isna(prev_ip): prev_ip = 'Unknown'

                    loc_ips.add(f"{cur_loc} ({cur_ip})")
                    loc_ips.add(f"{prev_loc} ({prev_ip})")
                    
                    if pd.notna(row['State']): states.add(row['State'])
                    if pd.notna(row['Prev_State']): states.add(row['Prev_State'])
                    
                if len(states) > 3:
                    high_risk_users.add(user)
                    
                loc_str = " / ".join(sorted(list(loc_ips)))
                user_blocks.append(f"{user} - {len(group)} instance(s) of impossible travel\nLocations: {loc_str}\n\n")

            if high_risk_users:
                travel_details.append("+" + "-" * 50 + "+\n")
                travel_details.append(f"| HIGH RISK: Users signing in from > 3 states{'':<6}|\n")
                travel_details.append("+" + "-" * 50 + "+\n")
                for h_user in sorted(list(high_risk_users)):
                    travel_details.append(f"| - {h_user:<46}|\n")
                travel_details.append("+" + "-" * 50 + "+\n\n")
                
            travel_details.extend(user_blocks)
            details.extend(travel_details)

        foreign = df[(~df['Country'].str.upper().isin(['US', 'USA', 'UNITED STATES', 'UNKNOWN'])) & (df['Country'].notna())].copy()
        if not foreign.empty:
            if 'Sign-in error code' in foreign.columns:
                foreign['Clean_Error'] = foreign['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
                foreign = foreign[foreign['Clean_Error'].isin(['0', 'None', 'nan', '', 'NaN'])]
            else:
                foreign['Clean_Error'] = "Unknown"

            if not foreign.empty:
                findings_found = True
                summary["Suspicious"].append(f"Successful Out-of-US Sign-ins: {len(foreign)} instance(s) detected.")
                details.append(f"[!] WARNING: Found {len(foreign)} SUCCESSFUL sign-in(s) originating outside the United States.\n")
                details.append("-" * 50 + "\n")
                for _, row in foreign.iterrows():
                    details.append(f"User: {row['User']}\n  Location: {row['Location']} (Detected Country: {row['Country']})\n  Time: {row.get('Date (UTC)', 'Unknown')} | IP: {row.get('IP address', 'Unknown')}\n  Error Code: {row.get('Clean_Error', 'Unknown')}\n\n")

    if cfg['errors']:
        df['Clean_Error'] = df['Sign-in error code'].astype(str).str.replace('.0', '', regex=False).str.strip()
        error_hits = df[df['Clean_Error'].isin(['50199', '500119', '90014'])]
        if not error_hits.empty:
            findings_found = True
            summary["Info"].append(f"Suspicious Errors: {len(error_hits)} instance(s) found.")
            details.append(f"[!] WARNING: Found {len(error_hits)} suspicious sign-in error code(s).\n")
            details.append("-" * 50 + "\n")
            for _, row in error_hits.iterrows():
                details.append(f"User: {row['User']} | Error Code: {row['Clean_Error']} | IP: {row.get('IP address', 'Unknown')} | Time: {row.get('Date (UTC)', 'Unknown')}\n")
            details.append("\n")

    if cfg['iocs']:
        leg_hits = df[df['Client app'].str.contains('Exchange ActiveSync|IMAP4|POP3|Legacy', na=False, case=False)]
        agt_hits = df[df['User agent'].str.contains('python-requests|curl|wget|nmap|powershell|httpclient', na=False, case=False)].copy()

        if not agt_hits.empty and 'Sign-in error code' in agt_hits.columns:
            agt_hits['Clean_Error'] = agt_hits['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
            agt_hits = agt_hits[agt_hits['Clean_Error'].isin(['0', 'None', 'nan', '', 'NaN'])]

        if not leg_hits.empty:
            findings_found = True
            summary["Info"].append(f"Legacy Auth: {len(leg_hits)} instance(s) found.")
            details.append(f"[!] WARNING: Found {len(leg_hits)} Legacy Auth attempts.\n")
            details.append("-" * 50 + "\n")
            for _, row in leg_hits.iterrows():
                details.append(f"User: {row['User']} | App: {row['Client app']} | IP: {row.get('IP address', 'Unknown')}\n")
            details.append("\n")
            
        if not agt_hits.empty:
            findings_found = True
            summary["Suspicious"].append(f"Suspicious User Agents: {len(agt_hits)} instance(s) found.")
            details.append(f"[!] WARNING: Found {len(agt_hits)} Suspicious Agent sign-ins (Successful Logins).\n")
            details.append("-" * 50 + "\n")
            for _, row in agt_hits.iterrows():
                details.append(f"User: {row['User']} | Agent: {row['User agent']} | IP: {row.get('IP address', 'Unknown')}\n")
            details.append("\n")

    return findings_found, summary, details

def analyze_non_interactive(df, log_type, params):
    findings_found = False
    summary = {"Critical": [], "Suspicious": [], "Info": []}
    details = []
    
    cfg = params['ni'] if log_type == "NonInteractive" else params['suni']
    
    req_cols = ['User']
    if cfg['agents']: req_cols.extend(['User agent', 'User type', 'Application'])
    if cfg['unmanaged']: req_cols.extend(['Compliant', 'Managed'])
    if cfg['refresh']: req_cols.append('Sign-in error code')
    if cfg['legacy']: req_cols.append('Client app')
    if cfg['travel']: req_cols.extend(['Date (UTC)', 'Location'])

    req_cols = list(set(req_cols))
    missing = [col for col in req_cols if col not in df.columns]
    if missing:
        summary["ERROR"] = f"Missing required columns:\n{', '.join(missing)}"
        return False, summary, details

    if log_type == "SUNI" and 'Location' in df.columns:
        findings_found = True
        loc_stats = df.dropna(subset=['Location']).groupby('Location').agg(
            count=('User', 'size'),
            unique_ips=('IP address', 'nunique') if 'IP address' in df.columns else ('User', lambda x: 0),
            ips=('IP address', lambda x: ', '.join(set(x.dropna().astype(str)))) if 'IP address' in df.columns else ('User', lambda x: 'Unknown')
        ).sort_values(by='count', ascending=False)
        
        summary["Info"].append("SUNI Sign-in Locations Profile generated.")
        details.append("--- SIGN-IN LOCATIONS SUMMARY (Most to Least Common) ---\n")
        for loc, row in loc_stats.iterrows():
            ip_warn = " [!] MULTIPLE IPs DETECTED" if row['unique_ips'] > 1 else ""
            details.append(f"Location: {loc} | Count: {row['count']} | Unique IPs: {row['unique_ips']}{ip_warn}\n")
            if row['unique_ips'] > 1:
                details.append(f"  IP Addresses: {row['ips']}\n")
        details.append("\n" + "-" * 50 + "\n")

    if cfg['travel']:
        df['Date (UTC)'] = pd.to_datetime(df['Date (UTC)'], errors='coerce')
        apply_vectorized_location(df)
        
        if 'Sign-in error code' in df.columns:
            df['Clean_Error_Travel'] = df['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
            valid_codes = ['0', 'None', 'nan', '', 'NaN', '50140', '50158', '50074', '50076', '50079', '50125']
            travel_df = df[df['Clean_Error_Travel'].isin(valid_codes)].copy()
        else:
            travel_df = df.copy()

        travel_df = travel_df.sort_values(['User', 'Date (UTC)']).dropna(subset=['Date (UTC)', 'State'])
        travel_df['Prev_State'] = travel_df.groupby('User')['State'].shift(1)
        travel_df['Prev_Location'] = travel_df.groupby('User')['Location'].shift(1)
        travel_df['Prev_Time'] = travel_df.groupby('User')['Date (UTC)'].shift(1)
        travel_df['Prev_IP'] = travel_df.groupby('User')['IP address'].shift(1) if 'IP address' in travel_df.columns else "Unknown"
        travel_df['Time_Diff_Mins'] = (travel_df['Date (UTC)'] - travel_df['Prev_Time']).dt.total_seconds() / 60.0
        
        imp_travel = travel_df[
            (travel_df['State'] != travel_df['Prev_State']) & 
            (travel_df['Prev_State'].notna()) & 
            (travel_df['Time_Diff_Mins'] < 30.0) &
            (travel_df['Time_Diff_Mins'] >= 0)
        ]

        if not imp_travel.empty:
            findings_found = True
            summary["Critical"].append(f"Impossible Travel (<30 mins): {len(imp_travel)} instance(s) detected.")
            
            travel_details = []
            travel_details.append(f"[!] WARNING: Found {len(imp_travel)} instance(s) of Impossible Travel (<30 mins between states).\n")
            travel_details.append("-" * 50 + "\n")
            
            high_risk_users = set()
            user_blocks = []
            
            for user, group in imp_travel.groupby('User'):
                loc_ips = set()
                states = set()
                
                for _, row in group.iterrows():
                    cur_loc = row['Location'] if pd.notna(row['Location']) else 'Unknown'
                    cur_ip = row.get('IP address', 'Unknown')
                    if pd.isna(cur_ip): cur_ip = 'Unknown'
                    
                    prev_loc = row['Prev_Location'] if pd.notna(row['Prev_Location']) else 'Unknown'
                    prev_ip = row.get('Prev_IP', 'Unknown')
                    if pd.isna(prev_ip): prev_ip = 'Unknown'

                    loc_ips.add(f"{cur_loc} ({cur_ip})")
                    loc_ips.add(f"{prev_loc} ({prev_ip})")
                    
                    if pd.notna(row['State']): states.add(row['State'])
                    if pd.notna(row['Prev_State']): states.add(row['Prev_State'])
                    
                if len(states) > 3:
                    high_risk_users.add(user)
                    
                loc_str = " / ".join(sorted(list(loc_ips)))
                user_blocks.append(f"{user} - {len(group)} instance(s) of impossible travel\nLocations: {loc_str}\n\n")

            if high_risk_users:
                travel_details.append("+" + "-" * 50 + "+\n")
                travel_details.append(f"| HIGH RISK: Users signing in from > 3 states{'':<6}|\n")
                travel_details.append("+" + "-" * 50 + "+\n")
                for h_user in sorted(list(high_risk_users)):
                    travel_details.append(f"| - {h_user:<46}|\n")
                travel_details.append("+" + "-" * 50 + "+\n\n")
                
            travel_details.extend(user_blocks)
            details.extend(travel_details)

        foreign = df[(~df['Country'].str.upper().isin(['US', 'USA', 'UNITED STATES', 'UNKNOWN'])) & (df['Country'].notna())].copy()
        if not foreign.empty:
            if 'Sign-in error code' in foreign.columns:
                foreign['Clean_Error'] = foreign['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
                foreign = foreign[foreign['Clean_Error'].isin(['0', 'None', 'nan', '', 'NaN'])]
            else:
                foreign['Clean_Error'] = "Unknown"

            if not foreign.empty:
                findings_found = True
                summary["Suspicious"].append(f"Successful Out-of-US Sign-ins: {len(foreign)} instance(s) detected.")
                details.append(f"[!] WARNING: Found {len(foreign)} SUCCESSFUL sign-in(s) originating outside the United States.\n")
                details.append("-" * 50 + "\n")
                for _, row in foreign.iterrows():
                    details.append(f"User: {row['User']}\n  Location: {row['Location']} (Detected Country: {row['Country']})\n  Time: {row.get('Date (UTC)', 'Unknown')} | IP: {row.get('IP address', 'Unknown')}\n  Error Code: {row.get('Clean_Error', 'Unknown')}\n\n")

    if cfg['agents']:
        member_df = df[df['User type'].astype(str).str.contains('Member', case=False, na=False)]
        agent_hits = member_df[member_df['User agent'].str.contains('python|curl|wget|nmap|powershell|httpclient|go-http-client', na=False, case=False)].copy()
        
        if not agent_hits.empty and 'Sign-in error code' in agent_hits.columns:
            agent_hits['Clean_Error'] = agent_hits['Sign-in error code'].fillna('0').astype(str).str.replace('.0', '', regex=False).str.strip()
            agent_hits = agent_hits[agent_hits['Clean_Error'].isin(['0', 'None', 'nan', '', 'NaN'])]

        if not agent_hits.empty:
            findings_found = True
            summary["Suspicious"].append(f"Scripting Agents on Standard Accounts: {len(agent_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(agent_hits)} instance(s) of Scripting Agents on Standard Member accounts (Successful Logins).\n")
            details.append("-" * 50 + "\n")
            for _, row in agent_hits.iterrows():
                details.append(f"User: {row['User']} | Agent: {row['User agent']} | App: {row.get('Application', 'Unknown')}\n")
            details.append("\n")

    if cfg['unmanaged']:
        unmanaged_hits = df[(df['Compliant'].astype(str).str.strip().str.lower() == 'no') & (df['Managed'].astype(str).str.strip().str.lower() == 'no')]
        if not unmanaged_hits.empty:
            findings_found = True
            summary["Critical"].append(f"Unmanaged/Non-Compliant Sign-ins: {len(unmanaged_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(unmanaged_hits)} sign-ins from Unmanaged & Non-Compliant devices.\n")
            details.append("-" * 50 + "\n")
            for user, count in unmanaged_hits['User'].value_counts().head(20).items():
                details.append(f"User: {user} | Total Unmanaged Hits: {count}\n")
            details.append("\n")

    if cfg['refresh']:
        df['Clean_Error'] = df['Sign-in error code'].astype(str).str.replace('.0', '', regex=False).str.strip()
        error_hits = df[df['Clean_Error'].isin(['50089', '50173'])]
        if not error_hits.empty:
            findings_found = True
            summary["Info"].append(f"Token Refresh/Revocation Errors: {len(error_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(error_hits)} Token Refresh/Revocation Errors.\n")
            details.append("-" * 50 + "\n")
            for _, row in error_hits.groupby(['User', 'Clean_Error']).size().reset_index(name='Count').iterrows():
                details.append(f"User: {row['User']} | Error Code: {row['Clean_Error']} | Count: {row['Count']}\n")
            details.append("\n")

    if cfg['legacy']:
        legacy_hits = df[df['Client app'].str.contains('Exchange ActiveSync|IMAP4|POP3|Legacy', na=False, case=False)]
        if not legacy_hits.empty:
            findings_found = True
            summary["Info"].append(f"Non-Interactive Legacy Auth: {len(legacy_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(legacy_hits)} Non-Interactive Legacy Auth attempts.\n")
            details.append("-" * 50 + "\n")
            for _, row in legacy_hits.iterrows():
                details.append(f"User: {row['User']} | App: {row['Client app']}\n")
            details.append("\n")

    return findings_found, summary, details

def analyze_purview(df, log_type, params):
    findings_found = False
    summary = {"Critical": [], "Suspicious": [], "Info": []}
    details = []
    cfg = params['purview']
    
    req_cols = ['UserId', 'Operation', 'CreationDate', 'AuditData']
    missing = [col for col in req_cols if col not in df.columns]
    if missing:
        summary["ERROR"] = f"Missing required columns for Purview checks:\n{', '.join(missing)}"
        return False, summary, details

    df['ParsedAudit'] = df['AuditData'].apply(safe_json_load)
    df['ClientIP'] = df['ParsedAudit'].apply(lambda x: x.get('ClientIPAddress', x.get('ClientIP', 'Unknown')))
    df['LogonType'] = df['ParsedAudit'].apply(lambda x: x.get('LogonType', 'Unknown'))
    df['MailboxOwnerUPN'] = df['ParsedAudit'].apply(lambda x: x.get('MailboxOwnerUPN', 'Unknown'))
    df['ClientProcessName'] = df['ParsedAudit'].apply(lambda x: x.get('ClientProcessName', x.get('ClientAppId', 'Unknown')))

    if cfg['inbox']:
        inbox_hits = df[df['Operation'].isin(['New-InboxRule', 'Set-InboxRule', 'UpdateInboxRules'])]
        if not inbox_hits.empty:
            findings_found = True
            summary["Suspicious"].append(f"Inbox Rule Creation/Modification: {len(inbox_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(inbox_hits)} Inbox Rule operations (Potential Exfiltration/Hiding).\n")
            details.append("-" * 50 + "\n")
            for _, row in inbox_hits.iterrows():
                params_list = row['ParsedAudit'].get('Parameters', [])
                rule_details = ", ".join([f"{p.get('Name')}={p.get('Value')}" for p in params_list if isinstance(p, dict)]) if isinstance(params_list, list) else str(params_list)
                details.append(f"Time: {row['CreationDate']} | User: {row['UserId']} | Op: {row['Operation']} | IP: {row['ClientIP']}\n  Details: {rule_details}\n\n")

    if cfg['mass_ops']:
        file_hits = df[df['Operation'].isin(['FileDeleted', 'HardDelete', 'FileDownloaded', 'SoftDelete', 'MoveToDeletedItems'])]
        if not file_hits.empty:
            op_counts = file_hits.groupby(['UserId', 'ClientIP', 'Operation']).size().reset_index(name='Count')
            suspicious_bulk = op_counts[op_counts['Count'] > 20]
            if not suspicious_bulk.empty:
                findings_found = True
                summary["Critical"].append(f"Mass File Operations (Downloads/Deletions): {len(suspicious_bulk)} bulk cluster(s).")
                details.append(f"[!] WARNING: Found bulk file/email operations (>20 occurrences per user/IP).\n")
                details.append("-" * 50 + "\n")
                for _, row in suspicious_bulk.iterrows():
                    details.append(f"User: {row['UserId']} | IP: {row['ClientIP']} | Operation: {row['Operation']} | Count: {row['Count']}\n")
                details.append("\n")

    if cfg['edisc']:
        edisc_hits = df[df['Operation'].str.contains('SearchQueryInitiated|eDiscoverySearchStarted|eDiscoveryExport|Search-Mailbox', case=False, na=False)]
        if not edisc_hits.empty:
            findings_found = True
            summary["Suspicious"].append(f"eDiscovery / Admin Search Operations: {len(edisc_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(edisc_hits)} eDiscovery/Search operations (Potential internal recon).\n")
            details.append("-" * 50 + "\n")
            for _, row in edisc_hits.iterrows():
                details.append(f"Time: {row['CreationDate']} | User: {row['UserId']} | Op: {row['Operation']} | IP: {row['ClientIP']}\n")
            details.append("\n")

    if cfg['mailbox']:
        mailbox_hits = df[(df['Operation'].isin(['MailItemsAccessed', 'MailboxLogin', 'FolderBind'])) & (df['LogonType'].isin(['Admin', 'Delegate']))]
        if not mailbox_hits.empty:
            findings_found = True
            summary["Critical"].append(f"Delegate/Admin Mailbox Access: {len(mailbox_hits)} instance(s).")
            details.append(f"[!] WARNING: Found {len(mailbox_hits)} Admin or Delegate Mailbox Access operations (Potential BEC).\n")
            details.append("-" * 50 + "\n")
            for _, row in mailbox_hits.iterrows():
                details.append(f"Time: {row['CreationDate']} | Actor: {row['UserId']} | Target: {row['MailboxOwnerUPN']} | Op: {row['Operation']} | LogonType: {row['LogonType']} | IP: {row['ClientIP']} | Client: {row['ClientProcessName']}\n")
            details.append("\n")

    return findings_found, summary, details

def browse_file():
    filename = filedialog.askopenfilename(title="Select Log File", filetypes=(("Log files", "*.csv *.xlsx *.xls"), ("All files", "*.*")))
    if filename: file_path_var.set(filename)

app = tk.Tk()
app.title("SOC Threat Hunt Analyzer")
app.geometry("550x700")
app.resizable(False, False)

file_path_var = tk.StringVar()
log_type_var = tk.StringVar(value="Interactive")

ip_filter_var = tk.BooleanVar(value=True); travel_filter_var = tk.BooleanVar(value=True); errors_filter_var = tk.BooleanVar(value=True); iocs_filter_var = tk.BooleanVar(value=True)
ni_travel_var = tk.BooleanVar(value=True); ni_agents_var = tk.BooleanVar(value=True); ni_unmanaged_var = tk.BooleanVar(value=True); ni_refresh_var = tk.BooleanVar(value=True); ni_legacy_var = tk.BooleanVar(value=True)
sui_travel_var = tk.BooleanVar(value=True); sui_errors_var = tk.BooleanVar(value=True); sui_iocs_var = tk.BooleanVar(value=True)
suni_travel_var = tk.BooleanVar(value=True); suni_agents_var = tk.BooleanVar(value=True); suni_unmanaged_var = tk.BooleanVar(value=True); suni_refresh_var = tk.BooleanVar(value=True); suni_legacy_var = tk.BooleanVar(value=True)
purview_inbox_var = tk.BooleanVar(value=True); purview_mass_op_var = tk.BooleanVar(value=True); purview_edisc_var = tk.BooleanVar(value=True); purview_mailbox_var = tk.BooleanVar(value=True)

tk.Label(app, text="1. Select Logs (.csv or .xlsx):", font=("Arial", 10, "bold")).pack(pady=(15, 5))
file_frame = tk.Frame(app); file_frame.pack(fill="x", padx=20)
tk.Entry(file_frame, textvariable=file_path_var, width=55, state='readonly').pack(side="left", padx=(0, 10))
tk.Button(file_frame, text="Browse...", command=browse_file, width=10).pack(side="left")

tk.Label(app, text="2. Select Log Dataset Type:", font=("Arial", 10, "bold")).pack(pady=(15, 5))
radio_container = tk.Frame(app); radio_container.pack(fill="x", padx=20)
row1 = tk.Frame(radio_container); row1.pack(anchor="center", pady=2)
tk.Radiobutton(row1, text="Interactive", variable=log_type_var, value="Interactive", command=toggle_options).pack(side="left", padx=10)
tk.Radiobutton(row1, text="Non-Interactive", variable=log_type_var, value="NonInteractive", command=toggle_options).pack(side="left", padx=10)
row2 = tk.Frame(radio_container); row2.pack(anchor="center", pady=2)
tk.Radiobutton(row2, text="Single User Int.", variable=log_type_var, value="SUI", command=toggle_options).pack(side="left", padx=10)
tk.Radiobutton(row2, text="Single User Non-Int.", variable=log_type_var, value="SUNI", command=toggle_options).pack(side="left", padx=10)
tk.Radiobutton(row2, text="Purview", variable=log_type_var, value="Purview", command=toggle_options).pack(side="left", padx=10)
row3 = tk.Frame(radio_container); row3.pack(anchor="center", pady=2)
tk.Radiobutton(row3, text="SessionID Color Sort", variable=log_type_var, value="ColorSort", command=toggle_options).pack(side="left", padx=10)

options_container = tk.Frame(app); options_container.pack(fill="x", padx=20, pady=10)
interactive_frame = tk.LabelFrame(options_container, text="Interactive Parameters", padx=10, pady=10)
tk.Checkbutton(interactive_frame, text="Shared Sessions: Require multiple unique IP addresses", variable=ip_filter_var).pack(anchor="w")
tk.Checkbutton(interactive_frame, text="Location Checks: Impossible Travel & Out-of-US Sign-ins", variable=travel_filter_var).pack(anchor="w")
tk.Checkbutton(interactive_frame, text="Suspicious Errors: MFA Denied, Max Age, Security Prompts", variable=errors_filter_var).pack(anchor="w")
tk.Checkbutton(interactive_frame, text="General IoCs: Legacy Auth & Suspicious User Agents", variable=iocs_filter_var).pack(anchor="w")

non_interactive_frame = tk.LabelFrame(options_container, text="Non-Interactive Parameters", padx=10, pady=10)
tk.Checkbutton(non_interactive_frame, text="Location Checks: Impossible Travel & Out-of-US Sign-ins", variable=ni_travel_var).pack(anchor="w")
tk.Checkbutton(non_interactive_frame, text="Scripting Agents: Flag 'Member' accounts using curl/python", variable=ni_agents_var).pack(anchor="w")
tk.Checkbutton(non_interactive_frame, text="Token Replay: Flag sign-ins from Unmanaged/Non-Compliant", variable=ni_unmanaged_var).pack(anchor="w")
tk.Checkbutton(non_interactive_frame, text="Refresh Failures: High freq of revoked/expired token errors", variable=ni_refresh_var).pack(anchor="w")
tk.Checkbutton(non_interactive_frame, text="Legacy Protocols: Background POP3/IMAP4 sync attempts", variable=ni_legacy_var).pack(anchor="w")

sui_frame = tk.LabelFrame(options_container, text="Single User Interactive Parameters", padx=10, pady=10)
tk.Checkbutton(sui_frame, text="Location Checks: Profile Locations, Velocity & Out-of-US", variable=sui_travel_var).pack(anchor="w")
tk.Checkbutton(sui_frame, text="Suspicious Errors: MFA Denied, Max Age, Security Prompts", variable=sui_errors_var).pack(anchor="w")
tk.Checkbutton(sui_frame, text="General IoCs: Legacy Auth & Suspicious User Agents", variable=sui_iocs_var).pack(anchor="w")

suni_frame = tk.LabelFrame(options_container, text="Single User Non-Interactive Parameters", padx=10, pady=10)
tk.Checkbutton(suni_frame, text="Location Checks: Profile Locations, Velocity & Out-of-US", variable=suni_travel_var).pack(anchor="w")
tk.Checkbutton(suni_frame, text="Scripting Agents: Flag 'Member' accounts using curl/python", variable=suni_agents_var).pack(anchor="w")
tk.Checkbutton(suni_frame, text="Token Replay: Flag sign-ins from Unmanaged/Non-Compliant", variable=suni_unmanaged_var).pack(anchor="w")
tk.Checkbutton(suni_frame, text="Refresh Failures: High freq of revoked/expired token errors", variable=suni_refresh_var).pack(anchor="w")
tk.Checkbutton(suni_frame, text="Legacy Protocols: Background POP3/IMAP4 sync attempts", variable=suni_legacy_var).pack(anchor="w")

purview_frame = tk.LabelFrame(options_container, text="Purview Audit Parameters", padx=10, pady=10)
tk.Checkbutton(purview_frame, text="Inbox Rules: New/Modified Inbox Forwarding Rules", variable=purview_inbox_var).pack(anchor="w")
tk.Checkbutton(purview_frame, text="Mass Operations: Bulk File Deletions or Downloads (>20)", variable=purview_mass_op_var).pack(anchor="w")
tk.Checkbutton(purview_frame, text="Reconnaissance: eDiscovery / Admin Search Executed", variable=purview_edisc_var).pack(anchor="w")
tk.Checkbutton(purview_frame, text="Mailbox Access: MailItemsAccessed & Delegate Logins", variable=purview_mailbox_var).pack(anchor="w")

color_sort_frame = tk.LabelFrame(options_container, text="SessionID Color Sort", padx=10, pady=10)
tk.Label(color_sort_frame, text="This feature isolates the 'Session ID' column and highlights unique sessions\nwith random, light colors to visually group them. It preserves your original\nfile and outputs a brand new Excel (.xlsx) document to your directory.", justify="left").pack(anchor="w")

interactive_frame.pack(fill="x")

run_btn = tk.Button(app, text="Run Analysis", command=run_check, bg="#2b2b2b", fg="#4CAF50", font=("Arial", 10, "bold"), width=20, height=2)
run_btn.pack(pady=5)
status_label = tk.Label(app, text="Status: Waiting for input", fg="gray")
status_label.pack(side="bottom", pady=10)

if __name__ == "__main__":
    app.mainloop()
